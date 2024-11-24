from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill


def format_existing_excel(file_path_or_bytes_object, header_row = 1, cushion = 5, freeze_headers = True):
    
    assert isinstance(file_path_or_bytes_object, (str, BytesIO)), "file_path_or_bytes_object must be a string or BytesIO"
    # Load the workbook and iterate through sheets
    
    if isinstance(file_path_or_bytes_object, BytesIO):
        file_path_or_bytes_object.seek(0)
        workbook = load_workbook(filename=BytesIO(file_path_or_bytes_object.read()))
    else:
        workbook = load_workbook(file_path_or_bytes_object)

    # Define a light grey fill
    grey_fill = PatternFill(
        start_color='00BABABA', 
        end_color='00BABABA', 
        fill_type='solid'
    )
    # Define a light fill for zebra striping
    stripe_fill = PatternFill(
        start_color='00DBDBDB', 
        end_color='00DBDBDB', 
        fill_type='solid'
    )
    # Define a border for the table body
    table_body_border = Border(
        left=Side(border_style='thin', color='00AAAAAA'),
        right=Side(border_style='thin', color='00AAAAAA'),
        top=Side(border_style='thin', color='00AAAAAA'),
        bottom=Side(border_style='thin', color='00AAAAAA')
    )
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Apply zebra striping starting from the row after the header
        for i, row in enumerate(sheet.iter_rows(min_row=header_row+1), start=header_row+1):
            for cell in row:
                cell.border = table_body_border
                if i % 2 == 0:  # For even row numbers
                    cell.fill = stripe_fill

        if freeze_headers == True:
            # Freeze the row just below the header row
            freeze_cell = 'A' + str(header_row + 1)
            sheet.freeze_panes = freeze_cell

        # Apply formatting to the specified header row
        for cell in sheet[header_row]:  # Use the header_row parameter
            cell.font = Font(bold=True)
            cell.border = Border(
                top=Side(style='thin'), 
                bottom=Side(style='thin'),
                left=Side(style='thin'), 
                right=Side(style='thin')
            )
            cell.fill = grey_fill

        # Set the column widths based on max length in column
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells[header_row-1:])
            adjusted_width = max_length + cushion  # Add cushion for a little extra space
            sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        # Apply filters to the specified header row
        if sheet.max_row >= header_row:  # Check if the header_row is within the data range
            sheet.auto_filter.ref = f"{sheet.dimensions.split(':')[0]}:{sheet.dimensions.split(':')[1]}"

    # Save the workbook
    if isinstance(file_path_or_bytes_object, BytesIO):
        outstream = BytesIO()
        workbook.save(outstream)
        outstream.seek(0)
        return outstream
    else:
        workbook.save(file_path_or_bytes_object)
        return
        